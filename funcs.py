import openpyxl
import re
import os
import json
import pandas as pd
import numpy as np
from asammdf import Signal, MDF
from PyQt6.QtWidgets import QFileDialog, QMessageBox
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.styles.colors import Color
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from scipy.signal import correlate
from scipy.signal import savgol_filter
from scipy.integrate import cumtrapz
from scipy.signal import find_peaks
from scipy.io import loadmat


# 序列化
def serialize_to_json(data, file_name):

    with open(file_name, 'w') as json_file:
        json.dump(data, json_file)


# 反序列化
def deserialize_from_json(file_name):

    with open(file_name, 'r') as json_file:
        return json.load(json_file)


# Resampling function
def resample_signal(time, values, interval):
    new_time = np.arange(time[0], time[-1], interval)
    new_values = np.interp(new_time, time, values)
    return new_time, new_values


# Smoothing function
def smooth_signal(values, window_length, polyorder):
    return savgol_filter(values, window_length, polyorder)


# Normalization function
def normalize_signal(values):
    min_val = np.min(values)
    max_val = np.max(values)
    return (values - min_val) / (max_val - min_val)


# Derivation function
def derive_signal(time, values):
    return np.gradient(values, time)


# Threshold filtering function
def apply_threshold(values, threshold):
    return np.where(values > threshold, values, 0)


# Function to apply a sequence of processing functions to the signal data
def process_signal_data(time, values, functions):
    processed_time = np.array(time)
    processed_values = np.array(values)
    for function, args in functions:
        if function == resample_signal:
            processed_time, processed_values = function(
                processed_time, processed_values, *args)
        elif function in [smooth_signal, normalize_signal, apply_threshold]:
            processed_values = function(processed_values, *args)
        elif function == derive_signal:
            processed_values = function(processed_time, processed_values)
    return processed_time, processed_values


# 处理MDF中信号重复的问题
def retrieve_signal_data(mdf_file, signal_name):
    groups = mdf_file.groups
    candidate_signals = []
    for i, group in enumerate(groups):
        for j, channel in enumerate(group.channels):
            if channel.name == signal_name:
                # Extract the signal based on group and index
                signal = mdf_file.get(signal_name, group=i, index=j)
                # Calculate sampling rate of this signal
                if signal.samples.size > 1:  # Avoid division by zero
                    rate = (signal.samples.size - 1) / \
                        (signal.timestamps[-1] - signal.timestamps[0])
                else:
                    rate = 0
                candidate_signals.append((rate, signal))

    if candidate_signals:
        # Select the signal with the highest sampling rate or the first one if equal
        _, highest_rate_signal = max(candidate_signals, key=lambda x: x[0])
        return highest_rate_signal.timestamps, highest_rate_signal.samples
    else:
        print(f"Signal {signal_name} not found.")
        return None, None


# 导入MDF文件后激活并更新下拉框选项
def update_and_enable_widgets(ui, unique_signals):
    """
    Updates and enables ComboBoxes, LineEdits, and Labels based on the unique signals from an MDF file.
    Args:
        ui: The UI instance containing the widgets.
        unique_signals: A list of unique signal names to add to the ComboBoxes.
    """
    for i in range(1, 9):  # Assuming you have 8 sets of widgets
        combobox = getattr(ui, f'comboBox_{i}')
        lineedit = getattr(ui, f'lineEdit_{i}')
        label = getattr(ui, f'label_{i}')

        # Update and enable ComboBox
        combobox.clear()
        combobox.addItems(unique_signals)
        combobox.setEnabled(True)

        # Enable LineEdit and Label
        lineedit.setEnabled(True)
        label.setEnabled(True)

        # Enable groupBox of CSV load
        ui.groupBox.setEnabled(True)


# 导入CSV文件后更新下拉框选项
def update_and_enable_combobox(ui, signals):
    for i in range(9, 16):
        combobox = getattr(ui, f'comboBox_{i}')
        combobox.clear()
        combobox.addItems(signals)
        combobox.addItem('无')
        combobox.setCurrentText('无')


# Importing and Merging CSV Files
def load_and_merge_csvs():
    # Let the user select multiple CSV files
    file_paths, _ = QFileDialog.getOpenFileNames(
        None, "Select CSV files", "", "CSV Files (*.csv)")
    if not file_paths:
        return None, None

    # Read and concatenate the CSV files
    all_dfs = []
    # Ensure files are sorted; adjust as necessary
    for file_path in sorted(file_paths):
        df = pd.read_csv(file_path)
        # Convert 'Date' and 'Time' into a single datetime column for easier sorting/merging
        df['DateTime'] = pd.to_datetime(df['Date'] + ' ' + df['Time'])
        all_dfs.append(df)

    # Concatenate all dataframes and sort by datetime
    full_df = pd.concat(all_dfs).sort_values(
        by='DateTime').reset_index(drop=True)
    return full_df, file_paths


# Extracting Specific Signals and Generating Time Differences
def extract_signals(ui, full_df):
    # Assuming 'DateTime' is your timestamp column
    full_df['TimeDiff'] = (full_df['DateTime'] -
                           full_df['DateTime'].iloc[0]).dt.total_seconds()
    # Filter columns that contain specific substrings
    filtered_columns = [col for col in full_df.columns if any(
        sub in col for sub in ['Udc', 'Idc', 'P'])]
    filtered_df = full_df[filtered_columns + ['TimeDiff']]
    # You can store these in the class or return them for further processing
    ui.csv_data = filtered_df


# 从字符串中筛选数字
def pick_number_from_string(s):
    # Use regular expression to find the first occurrence of a number in the string
    match = re.search(r'\d+', s)
    if match:
        return match.group()  # Convert found number to integer
    return None  # Return None if no number is found


# Function to add 'TimeDiff' if not present in DataFrame
def add_timediff_if_missing(ui, data_frame):
    if all(['TimeDiff' not in item for item in data_frame.columns]):
        # Number of rows in the DataFrame
        num_rows = len(data_frame)
        # Create TimeDiff series starting from 0, incrementing by xHz up to the number of rows
        selected_action = ui.freqGrope.checkedAction()
        freq = int(pick_number_from_string(selected_action.text()))
        time_sample = 1 / freq
        time_diff_series = np.arange(0, num_rows * time_sample, time_sample)
        # Adjust the length in case the initial array is longer than the dataframe due to floating point range generation
        time_diff_series = time_diff_series[:num_rows]
        # Add this series to the DataFrame
        data_frame['TimeDiff'] = time_diff_series
    return data_frame


# 从.log文件中读取转鼓数据
def read_log_file(ui, file_name):
    with open(file_name, 'r') as file:
        lines = file.readlines()
        # Find the indexes for Names, DataStart, and DataEnd
        names_index = lines.index('[Names]\n') + 1
        data_start_index = lines.index('[DataStart]\n') + 1
        data_end_index = lines.index('[DataEnd]\n')
        # Extract header and units, then combine them
        headers = lines[names_index].strip().split(';')
        units = lines[names_index + 1].strip().split(';')
        combined_headers = [f"{h} {u}" for h, u in zip(headers, units)]
        # Extract data
        data = [line.strip().split(';')
                for line in lines[data_start_index:data_end_index]]
        # Convert to DataFrame
        data = pd.DataFrame(data, columns=combined_headers)
        data = data.drop(data.columns[-1], axis=1)
        data = add_timediff_if_missing(ui, data)
        return data


# 从.txt文件中读取转鼓数据
def read_txt_file(ui, file_name):
    with open(file_name, 'r') as file:
        lines = file.readlines()
        # Extract header and units, then combine them
        headers = lines[0].strip().split('\t')
        units = lines[1].strip().split('\t')
        combined_headers = [f"{h} [{u}]" for h, u in zip(headers, units)]
        # Extract data
        data = [line.strip().split('\t')
                for line in lines[2:]]  # From third line onwards
        # Convert to DataFrame
        data = pd.DataFrame(data, columns=combined_headers)
        data = add_timediff_if_missing(ui, data)
        return data


# 从.csv文件中读取转鼓数据
def read_csv_file(ui, file_name):
    data = pd.read_csv(file_name)
    data = add_timediff_if_missing(ui, data)
    return data


# 警告提示对话框
def show_warning(message):
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Icon.Warning)
    msg_box.setText(message)
    msg_box.setWindowTitle("Warning")
    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
    msg_box.exec()


# 消息提示对话框
def show_must_read_info(message):
    msg_box = QMessageBox()
    msg_box.setIcon(QMessageBox.Icon.Information)
    msg_box.setWindowTitle("提示")
    msg_box.setText(message)
    msg_box.setStandardButtons(QMessageBox.StandardButton.Ok)
    msg_box.exec()


# 分循环积分
def power_integration(power, start_time, end_time, flage):
    # Initialize an empty list to store the integration results for each segment
    integration_results = []

    for start, end in zip(start_time, end_time):
        # Filter the power for the current start and end time range
        filtered_power = power[(power[:, 0] >= start) & (power[:, 0] <= end)]
        filtered_flage = flage[(flage[:, 0] >= start) & (flage[:, 0] <= end)]

        if len(filtered_power) != len(filtered_flage):
            filtered_power_interp = np.interp(
                filtered_flage[:, 0], filtered_power[:, 0], filtered_power[:, 1], 0, 0)
            filtered_power = np.hstack(
                (filtered_flage[:, 0].reshape(-1, 1), filtered_power_interp.reshape(-1, 1)))

        # Initialize integration values for non-negative and negative parts
        integrated_power_positive = 0
        integrated_power_negative = 0
        # Compute the power integration for non-negative and negative parts separately
        integrated_power = np.trapz(
            filtered_power[:, 1], filtered_power[:, 0])
        # Non-negative part
        positive_power = np.where(
            filtered_flage[:, 1] > 0, filtered_power[:, 1], 0)
        integrated_power_positive = np.trapz(
            positive_power, filtered_power[:, 0])
        # Negative part
        negative_power = np.where(
            filtered_flage[:, 1] < 0, filtered_power[:, 1], 0)
        integrated_power_negative = np.trapz(
            negative_power, filtered_power[:, 0])

        # Convert power integration from watt-seconds to watt-hours for consistency
        integration_results.append(
            [integrated_power / 3600, integrated_power_positive / 3600, integrated_power_negative / 3600])

    # Convert the results list to a NumPy ndarray
    return np.array(integration_results)


# 通过互相关函数对齐信号
def align_signals_via_correlation(csv_data, mdf_data):
    # Extract the current signals and compute cross-correlation
    csv_current = csv_data[:, 1]
    mdf_current = mdf_data[:, 1]
    cross_corr = correlate(mdf_current, csv_current,
                           mode='same', method='auto')

    # Find the index of the maximum correlation to determine the shift
    shift_idx = np.argmax(cross_corr) - (len(csv_current) // 2)

    # Compute time shift based on sampling rate (assuming uniform sampling)
    time_shift = shift_idx * np.mean(np.diff(csv_data[:, 0]))

    return time_shift


# 互相关性检查
def correlation_check(data1, data2, start_time, end_time):
    correlation_results = []

    for start, end in zip(start_time, end_time):
        filtered_data1 = data1[(data1[:, 0] >= start) &
                               (data1[:, 0] <= end), 1]
        filtered_data2 = data2[(data2[:, 0] >= start) &
                               (data2[:, 0] <= end), 1]
        correlation = np.correlate(filtered_data1, filtered_data2, mode='same')
        min_mean = np.mean(np.partition(correlation, 100)[:100])
        correlation_results.append(
            [np.mean(correlation) / np.max(correlation), min_mean / np.mean(correlation)])

    # Convert the results list to a NumPy ndarray
    return np.array(correlation_results)


# 从lineEdit中读取数据并存储
def read_data_from_lineedit(ui, attribute_name_list):
    for i in range(9, 13):
        lineEdit = getattr(ui, f'lineEdit_{i}')
        if lineEdit.text():
            setattr(ui, attribute_name_list[i-9], float(lineEdit.text()))


# 创建Excel模板的详细信息
def create_excel_from_details(details, new_file_path):
    new_wb = openpyxl.Workbook()
    new_wb.remove(new_wb.active)  # Remove the default sheet

    for sheet_name, sheet_details in details.items():
        new_sheet = new_wb.create_sheet(title=sheet_name)
        new_sheet.sheet_properties.tabColor = sheet_details['sheet_properties'].get(
            'tab_color', None)

        # Set row heights and column widths
        for row, height in sheet_details['row_heights'].items():
            if height:  # Ensure the height is not None
                new_sheet.row_dimensions[row].height = height
        for col, width in sheet_details['column_widths'].items():
            if width:  # Ensure the width is not None
                new_sheet.column_dimensions[col].width = width

        # Set cell values, styles, and merged cells
        for cell_address, cell_details in sheet_details['cells'].items():
            cell = new_sheet[cell_address]
            cell.value = cell_details['value']
            if cell_details['formula']:
                cell.value = "=" + cell_details['formula']

            # Check and set fill color
            start_color = cell_details['fill_color']['start_color']
            # end_color=cell_details['fill_color']['end_color']
            if isinstance(start_color, str):
                start_color = Color(rgb=start_color)
            else:
                # Default color (white) if no valid color provided
                start_color = Color(rgb='FFFFFFFF')
            cell.fill = PatternFill(
                start_color=start_color, end_color=start_color, fill_type=cell_details['fill_color']['fill_type'])
            cell.font = Font(
                sz=cell_details['font_size'], name=cell_details['font_name'])
            cell.border = Border(
                left=Side(style=cell_details['border']['left']),
                right=Side(style=cell_details['border']['right']),
                top=Side(style=cell_details['border']['top']),
                bottom=Side(style=cell_details['border']['bottom'])
            )
            cell.alignment = Alignment(
                horizontal=cell_details['alignment']['horizontal'],
                vertical=cell_details['alignment']['vertical'],
                wrap_text=cell_details['alignment']['wrap_text']
            )

        # Apply merged cells
        for merge_cells_range in sheet_details['merged_cells']:
            new_sheet.merge_cells(merge_cells_range)

    # Save the new workbook
    new_wb.save(new_file_path)


# 将N维数组写入Excel
def write_ndarray_to_excel(ndarray, sheet_index, start_cell, file_path):
    """
    Writes the contents of a numpy.ndarray to an Excel sheet and formats the cells
    to display two decimal places.

    :param ndarray: The numpy.ndarray object to write to the Excel sheet.
    :param sheet_index: Index of the sheet in the Excel workbook (0-based).
    :param start_cell: String representation of the starting cell (e.g., 'C2').
    :param file_path: Path to the Excel file to write to.
    """
    # Ensure ndarray is at least two-dimensional
    if ndarray.ndim == 1:
        ndarray = ndarray.reshape(-1, 1)  # Change shape from (n,) to (n, 1)

    # Load the workbook and select the specified sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.worksheets[sheet_index]

    # Determine the starting row and column from the start_cell parameter
    start_column_letter, start_row = coordinate_from_string(start_cell)
    start_column = column_index_from_string(start_column_letter)

    # Write the ndarray contents to the sheet and format cells
    for i in range(ndarray.shape[0]):
        for j in range(ndarray.shape[1]):
            cell = sheet.cell(row=start_row + i, column=start_column + j)
            cell.value = ndarray[i, j]
            cell.number_format = '0.00'

    # Save the updated workbook
    wb.save(file_path)


# 为功率分析仪数据生成信号列表
def generate_signals_for_csv(ui, attribute_name, time, filtered_time, start_time, end_time):
    if hasattr(ui, attribute_name):
        signals = []
        data = getattr(ui, attribute_name)
        curr = data['Idc'][(time >= start_time) & (time <= end_time)]
        volt = data['Udc'][(time >= start_time) & (time <= end_time)]
        power = data['P'][(time >= start_time) & (time <= end_time)]
        signal_curr = Signal(
            samples=np.array(curr, dtype=np.float64),
            timestamps=np.array(
                filtered_time, dtype=np.float64),
            name=attribute_name + '_curr [A]'
        )
        signal_volt = Signal(
            samples=np.array(volt, dtype=np.float64),
            timestamps=np.array(
                filtered_time, dtype=np.float64),
            name=attribute_name + '_volt [V]'
        )
        signal_power = Signal(
            samples=np.array(power, dtype=np.float64),
            timestamps=np.array(
                filtered_time, dtype=np.float64),
            name=attribute_name + '_volt [kW]'
        )
        signals.append(signal_curr)
        signals.append(signal_volt)
        signals.append(signal_power)
        return signals
    else:
        return None


# 数据分析，run action回调函数的核心程序
def data_analysis(ui, progressDialog):
    # 处理转鼓数据
    if hasattr(ui, 'dyno_vehspd') or hasattr(ui, 'dyno_vehacc'):
        if not hasattr(ui, 'dyno_vehspd'):
            if hasattr(ui, 'dyno_vehacc'):
                vehacc = ui.dyno_vehacc
                vehspd_intg = cumtrapz(
                    vehacc[:, 1], vehacc[:, 0], initial=0) * 3.6
                ui.dyno_vehspd = np.hstack(
                    (vehacc[:, 0].reshape(-1, 1), vehspd_intg.reshape(-1, 1)))
                ui.cycle_start_time_dyno, ui.cycle_end_time_dyno = cycle_divide(
                    ui, 'dyno_vehspd')
                vehspd_intg_list = []
                for start_time, end_time in zip(ui.start_time_dyno[1:], ui.end_time_dyno[1:]):
                    filtered_time = vehacc[(vehacc[:, 0] >= start_time) & (
                        vehacc[:, 0] <= end_time), 0]
                    filtered_value = vehacc[(vehacc[:, 0] >= start_time) & (
                        vehacc[:, 0] <= end_time), 1]
                    vehspd_intg_cycle = cumtrapz(
                        filtered_value, filtered_time, initial=0) * 3.6
                    vehspd_intg_list.append(
                        np.hstack((filtered_time.reshape(-1, 1), vehspd_intg_cycle.reshape(-1, 1))))
                combined_array = np.vstack(vehspd_intg_list)
                sorted_array = combined_array[combined_array[:, 0].argsort()]
                _, unique_indices = np.unique(
                    sorted_array[:, 0], return_index=True)
                unique_sorted_array = sorted_array[unique_indices]
                vehspd_intg = np.interp(
                    vehacc[:, 0], unique_sorted_array[:, 0], unique_sorted_array[:, 1], 0, 0)
                ui.dyno_data['vehspd integral [km/h]'] = vehspd_intg
                ui.dyno_vehspd = np.hstack(
                    (vehacc[:, 0].reshape(-1, 1), vehspd_intg.reshape(-1, 1)))
        else:
            ui.cycle_start_time_dyno, ui.cycle_end_time_dyno = cycle_divide(
                ui, 'dyno_vehspd')

        # 手动增加循环
        if hasattr(ui, 'add_cycle_info'):
            time_shift = np.mean(
                np.array(ui.cycle_start_time_dyno) - np.array(ui.cycle_start_time))
            data = ui.add_cycle_info
            add_cycle_start_time = data[0] + time_shift
            add_cycle_end_time = data[1] + time_shift
            loc = int(pick_number_from_string(data[2])) - 1
            if data[3] == '后':
                loc = loc + 1
            ui.cycle_start_time_dyno.insert(loc, add_cycle_start_time)
            ui.cycle_end_time_dyno.insert(loc, add_cycle_end_time)
        # 手动删除循环
        if hasattr(ui, 'del_cycle_info'):
            loc = int(pick_number_from_string(ui.del_cycle_info)) - 1
            ui.cycle_start_time_dyno.pop(loc)
            ui.cycle_end_time_dyno.pop(loc)
        # 通过循环工况时间计算全工况时间
        ui.start_time_dyno, ui.end_time_dyno = cycle_time_to_start_time(
            ui, ui.cycle_start_time_dyno, ui.cycle_end_time_dyno, ui.dyno_vehspd[len(ui.dyno_vehspd) - 1, 0])
        # 积分运算
        ui.dyno_distance = power_integration(
            ui.dyno_vehspd, ui.start_time_dyno, ui.end_time_dyno, ui.dyno_vehspd)[:, 1]
        progressDialog.setValue(15)

        attribute_name_list = ['F0', 'F1', 'F2', 'equ_mass']
        read_data_from_lineedit(ui, attribute_name_list)
        if hasattr(ui, 'F0') and hasattr(ui, 'F1') and hasattr(ui, 'F2') and hasattr(ui, 'equ_mass') and hasattr(ui, 'dyno_vehacc'):
            coefficients = [ui.F2, ui.F1, ui.F0]
            poly_func = np.poly1d(coefficients)
            drive_shaft_force = poly_func(
                ui.dyno_vehspd[:, 1]) + ui.equ_mass * ui.dyno_vehacc[:, 1]
            drive_shaft_power = drive_shaft_force * ui.dyno_vehspd[:, 1] / 3600
            kinetic_power = ui.equ_mass * \
                ui.dyno_vehacc[:, 1] * ui.dyno_vehspd[:, 1] / 3600
            resistance_power = poly_func(
                ui.dyno_vehspd[:, 1]) * ui.dyno_vehspd[:, 1] / 3600
            ui.drive_shaft_power = np.hstack(
                (ui.dyno_vehacc[:, 0].reshape(-1, 1), drive_shaft_power.reshape(-1, 1)))
            ui.kinetic_power = np.hstack(
                (ui.dyno_vehacc[:, 0].reshape(-1, 1), kinetic_power.reshape(-1, 1)))
            ui.resistance_power = np.hstack(
                (ui.dyno_vehacc[:, 0].reshape(-1, 1), resistance_power.reshape(-1, 1)))
            ui.drive_shaft_energy = power_integration(
                ui.drive_shaft_power, ui.start_time_dyno, ui.end_time_dyno, ui.drive_shaft_power)
            ui.kinetic_energy = power_integration(
                ui.kinetic_power, ui.start_time_dyno, ui.end_time_dyno, ui.drive_shaft_power)
            ui.resistance_energy = power_integration(
                ui.resistance_power, ui.start_time_dyno, ui.end_time_dyno, ui.drive_shaft_power)
            progressDialog.setValue(25)

            kinetic_power_std = ui.equ_mass * \
                ui.std_cycle['vehacc'] * ui.std_cycle['vehspd_trgt'] / 3600
            resistance_power_std = poly_func(
                ui.std_cycle['vehspd_trgt']) * ui.std_cycle['vehspd_trgt'] / 3600
            drive_shaft_power_std = kinetic_power_std + resistance_power_std
            ui.drive_shaft_power_std = np.hstack(
                (ui.std_cycle['time'].reshape(-1, 1), drive_shaft_power_std.reshape(-1, 1)))
            ui.drive_shaft_energy_std = power_integration(
                ui.drive_shaft_power_std, [1], [1800], ui.drive_shaft_power_std)
            ui.kinetic_power_std = np.hstack(
                (ui.std_cycle['time'].reshape(-1, 1), kinetic_power_std.reshape(-1, 1)))
            ui.kinetic_energy_std = power_integration(
                ui.kinetic_power_std, [1], [1800], ui.drive_shaft_power_std)
            ui.resistance_power_std = np.hstack(
                (ui.std_cycle['time'].reshape(-1, 1), resistance_power_std.reshape(-1, 1)))
            ui.resistance_energy_std = power_integration(
                ui.resistance_power_std, [1], [1800], ui.drive_shaft_power_std)
            progressDialog.setValue(30)

    # 处理MDF数据和功率分析仪数据
    if hasattr(ui, 'mdf_vehspd'):
        ui.cycle_start_time, ui.cycle_end_time = cycle_divide(ui, 'mdf_vehspd')
        ui.menu_10.setEnabled(True)
        ui.menu_11.setEnabled(True)
        if hasattr(ui, 'add_cycle_info'):
            data = ui.add_cycle_info
            loc = int(pick_number_from_string(data[2])) - 1
            if data[3] == '后':
                loc = loc + 1
            ui.cycle_start_time.insert(loc, data[0])
            ui.cycle_end_time.insert(loc, data[1])
        if hasattr(ui, 'del_cycle_info'):
            loc = int(pick_number_from_string(ui.del_cycle_info)) - 1
            ui.cycle_start_time.pop(loc)
            ui.cycle_end_time.pop(loc)

        ui.start_time, ui.end_time = cycle_time_to_start_time(
            ui, ui.cycle_start_time, ui.cycle_end_time, ui.mdf_vehspd[len(ui.mdf_vehspd) - 1, 0])
        ui.mdf_distance = power_integration(
            ui.mdf_vehspd, ui.start_time, ui.end_time, ui.mdf_vehspd)[:, 1]
        progressDialog.setValue(40)

        if hasattr(ui, 'mdf_battcurr') and hasattr(ui, 'mdf_battvolt'):
            ui.mdf_battpower = np.zeros_like(ui.mdf_battcurr)
            ui.mdf_battpower[:, 0] = ui.mdf_battcurr[:, 0]
            ui.mdf_battpower[:, 1] = ui.mdf_battcurr[:, 1] * \
                ui.mdf_battvolt[:, 1] / 1000
            ui.mdf_battenergy = power_integration(
                ui.mdf_battpower, ui.start_time, ui.end_time, ui.mdf_battpower)
            progressDialog.setValue(51)

        if hasattr(ui, 'mdf_dcdccurr') and hasattr(ui, 'mdf_dcdcvolt'):
            ui.mdf_dcdcpower = np.zeros_like(ui.mdf_dcdccurr)
            ui.mdf_dcdcpower[:, 0] = ui.mdf_dcdccurr[:, 0]
            ui.mdf_dcdcpower[:, 1] = ui.mdf_dcdccurr[:, 1] * \
                ui.mdf_dcdcvolt[:, 1] / 1000
            ui.mdf_dcdcenergy = power_integration(
                ui.mdf_dcdcpower, ui.start_time, ui.end_time, ui.mdf_battpower)
            progressDialog.setValue(62)

        if hasattr(ui, 'mdf_ecppwr'):
            mdf_ecppwr = np.zeros_like(ui.mdf_ecppwr)
            mdf_ecppwr[:, 0] = ui.mdf_ecppwr[:, 0]
            mdf_ecppwr[:, 1] = ui.mdf_ecppwr[:, 1] / 1000
            ui.mdf_ecpenergy = power_integration(
                mdf_ecppwr, ui.start_time, ui.end_time, ui.mdf_battpower)
            progressDialog.setValue(73)

        if hasattr(ui, 'mdf_ptcpwr'):
            mdf_ptcpwr = np.zeros_like(ui.mdf_ptcpwr)
            mdf_ptcpwr[:, 0] = ui.mdf_ptcpwr[:, 0]
            mdf_ptcpwr[:, 1] = ui.mdf_ptcpwr[:, 1] / 1000
            ui.mdf_ptcenergy = power_integration(
                mdf_ptcpwr, ui.start_time, ui.end_time, ui.mdf_battpower)

        progressDialog.setValue(84)

        if not hasattr(ui, 'csv_batt'):
            if hasattr(ui, 'csv_ipu') and hasattr(ui, 'csv_ips'):
                time = ui.csv_ipu['TimeDiff']
                csv_batt_Idc = ui.csv_ipu['Idc'] + ui.csv_ips['Idc']
                csv_batt_P = ui.csv_ipu['P'] + ui.csv_ips['P']
                ui.csv_batt = pd.DataFrame({
                    'Idc': csv_batt_Idc,
                    'Udc': ui.csv_ipu['P'],
                    'P': csv_batt_P,
                    'TimeDiff': time
                })
            else:
                progressDialog.setValue(100)
                show_must_read_info("目前导入的数据已完成分析，可进行输出！")
                return
        else:
            time = ui.csv_batt['TimeDiff']
            csv_batt_Idc = ui.csv_batt['Idc']
            csv_batt_P = ui.csv_batt['P']
        ui.csv_battcurr = np.hstack(
            (time.to_numpy().reshape(-1, 1), csv_batt_Idc.to_numpy().reshape(-1, 1)))
        if hasattr(ui, 'mdf_battcurr'):
            resample_batcurr = np.interp(
                ui.csv_battcurr[:, 0], ui.mdf_battcurr[:, 0], ui.mdf_battcurr[:, 1], 0, 0)
            batcurr_resample = np.hstack(
                (ui.csv_battcurr[:, 0].reshape(-1, 1), resample_batcurr.reshape(-1, 1)))
            time_shift = align_signals_via_correlation(
                ui.csv_battcurr, batcurr_resample)
            # Apply time shift:
            # If time_shift is positive, MDF starts later; we shift MDF time backwards.
            # If time_shift is negative, CSV starts later; we shift CSV time backwards.
            if time_shift > 0:
                ui.aligned_csv_time = ui.csv_battcurr[:, 0] + time_shift
                ui.start_time_csv = ui.start_time
                ui.end_time_csv = ui.end_time
                ui.cycle_start_time_csv = ui.cycle_start_time
                ui.cycle_end_time_csv = ui.cycle_end_time
            else:
                ui.aligned_csv_time = ui.csv_battcurr[:, 0]
                ui.start_time_csv = ui.start_time - time_shift
                ui.end_time_csv = ui.end_time - time_shift
                ui.cycle_start_time_csv = ui.cycle_start_time - time_shift
                ui.cycle_end_time_csv = ui.cycle_end_time - time_shift
            ui.csv_battpower = np.hstack(
                (ui.aligned_csv_time.reshape(-1, 1), csv_batt_P.to_numpy().reshape(-1, 1) / 1000))
            ui.csv_battenergy = power_integration(
                ui.csv_battpower, ui.start_time_csv, ui.end_time_csv, ui.csv_battpower)
            progressDialog.setValue(92)

            if hasattr(ui, 'csv_ipu'):
                ui.csv_ipupower = np.hstack(
                    (ui.aligned_csv_time.reshape(-1, 1), ui.csv_ipu['P'].to_numpy().reshape(-1, 1) / 1000))
                ui.csv_ipuenergy = power_integration(
                    ui.csv_ipupower, ui.start_time_csv, ui.end_time_csv, ui.csv_battpower)
            if hasattr(ui, 'csv_ips'):
                ui.csv_ipspower = np.hstack(
                    (ui.aligned_csv_time.reshape(-1, 1), ui.csv_ips['P'].to_numpy().reshape(-1, 1) / 1000))
                ui.csv_ipsenergy = power_integration(
                    ui.csv_ipspower, ui.start_time_csv, ui.end_time_csv, ui.csv_battpower)
            if hasattr(ui, 'csv_dcdc_high'):
                ui.csv_dcdchigh_power = np.hstack(
                    (ui.aligned_csv_time.reshape(-1, 1), ui.csv_dcdc_high['P'].to_numpy().reshape(-1, 1) / 1000))
                ui.csv_dcdchigh_energy = power_integration(
                    ui.csv_dcdchigh_power, ui.start_time_csv, ui.end_time_csv, ui.csv_battpower)

            if hasattr(ui, 'csv_dcdc_low'):
                ui.csv_dcdclow_power = np.hstack(
                    (ui.aligned_csv_time.reshape(-1, 1), ui.csv_dcdc_low['P'].to_numpy().reshape(-1, 1) / 1000))
                ui.csv_dcdclow_energy = power_integration(
                    ui.csv_dcdclow_power, ui.start_time_csv, ui.end_time_csv, ui.csv_battpower)

            if hasattr(ui, 'csv_ecp'):
                ui.csv_ecppower = np.hstack(
                    (ui.aligned_csv_time.reshape(-1, 1), ui.csv_ecp['P'].to_numpy().reshape(-1, 1) / 1000))
                ui.csv_ecpenergy = power_integration(
                    ui.csv_ecppower, ui.start_time_csv, ui.end_time_csv, ui.csv_battpower)

            if hasattr(ui, 'csv_ptc'):
                ui.csv_ptcpower = np.hstack(
                    (ui.aligned_csv_time.reshape(-1, 1), ui.csv_ptc['P'].to_numpy().reshape(-1, 1) / 1000))
                ui.csv_ptcenergy = power_integration(
                    ui.csv_ptcpower, ui.start_time_csv, ui.end_time_csv, ui.csv_battpower)
            progressDialog.setValue(96)
    progressDialog.setValue(100)
    show_must_read_info("目前导入的数据已完成分析，可进行输出！")


# 数据归一化
def normalize_signal(signal):
    """Normalize signal to zero mean and unit variance."""
    signal_mean = np.mean(signal)
    signal_std = np.std(signal)
    return (signal - signal_mean) / signal_std


# 在车速中识别标准工况
def find_std_in_vehspd(std, vehspd, coef):
    """Find instances of 'std' in 'vehspd' using cross-correlation."""
    # Extract speed values and normalize
    std_speed = normalize_signal(std[:, 1])
    vehspd_speed = normalize_signal(vehspd[:, 1])

    # Cross-correlate
    correlation = np.correlate(vehspd_speed, std_speed, mode='full')

    # Find peaks in the cross-correlation that could indicate matches
    peaks, _ = find_peaks(correlation, height=np.max(
        correlation)*coef)  # Example threshold

    # Convert peak indices to time values from 'vehspd'
    # Note: Adjusting for the length of std to align with the start of the matching segment
    peak_indices = peaks - len(std_speed) + 1
    peak_times = vehspd[peak_indices, 0] if len(
        vehspd) > 0 and len(peak_indices) > 0 else []

    return peak_times, peak_indices


# 划分工况
def cycle_divide(ui, attribute_name):
    vehspd = getattr(ui, attribute_name)
    # Load the .mat file
    selected_cycle = ui.cycleGrope.checkedAction()
    std_cycle_file = './data/' + selected_cycle.text() + '.mat'
    data = loadmat(std_cycle_file)
    ui.std_cycle = data
    std = np.hstack((data['time'], data['vehspd_trgt']))
    resample_time, resample_values = resample_signal(
        vehspd[:, 0], vehspd[:, 1], 1)
    vehspd_resample = np.hstack((resample_time.reshape(-1, 1),
                                resample_values.reshape(-1, 1)))
    peak_times, _ = find_std_in_vehspd(
        std, vehspd_resample, ui.correlation_coef)
    # print("Potential match times:", peak_times)
    # Loop through each peak index
    cycle_start_time = []
    cycle_end_time = []
    for _, time in enumerate(peak_times):
        cycle_start_index = np.searchsorted(vehspd[:, 0], time)
        cycle_start_time.append(vehspd[cycle_start_index, 0])
        cycle_end_index = np.searchsorted(vehspd[:, 0], time+1799)
        cycle_end_time.append(
            vehspd[min(cycle_end_index, len(vehspd) - 1), 0])
    return cycle_start_time, cycle_end_time


# 根据循环起始和结束时间生成总的起始和结束时间
def cycle_time_to_start_time(ui, cycle_start_time, cycle_end_time, data_end):
    selected_type = ui.actionGroup.checkedAction()
    if selected_type.text() == '缩短法':
        if len(cycle_start_time) == 4:
            start_time = [cycle_start_time[0]] + cycle_start_time + \
                [cycle_end_time[1], cycle_end_time[3]]
            end_time = [data_end] + cycle_end_time + \
                [cycle_start_time[2], data_end]
        else:
            start_time = [cycle_start_time[0]] + cycle_start_time
            end_time = [cycle_end_time[-1]] + cycle_end_time
            show_must_read_info('工况数异常，请确认数据是否完整，或在设置中调整工况相关性系数!')
    else:
        start_time = [cycle_start_time[0]] + cycle_start_time
        end_time = [cycle_end_time[-1]] + cycle_end_time
    return start_time, end_time


# 分循环保存LOG数据
def loop_for_save_log(ui, progressDialog, original_mdf, directory):
    num_steps = len(ui.cycle_start_time)
    for i, (start_time, end_time) in enumerate(zip(ui.cycle_start_time, ui.cycle_end_time)):
        progressDialog.setValue(round(((i) / num_steps) * 100))
        # Extract the segment from the original MDF file
        segment = original_mdf.cut(
            start=start_time, stop=end_time, time_from_zero=True)

        # Save the segment to a new MDF file
        segment_file_name = f"LOG_Cycle_{i+1}.mf4"
        segment_file_path = os.path.join(directory, segment_file_name)
        segment.save(segment_file_path)
    progressDialog.setValue(100)
    show_must_read_info("已将LOG数据生成分循环的MDF数据!")


# 分循环保存功率分析仪数据
def loop_for_save_power(ui, progressDialog, directory):

    num_steps = len(ui.cycle_start_time_csv)
    for i, (start_time, end_time) in enumerate(zip(ui.cycle_start_time_csv, ui.cycle_end_time_csv)):
        progressDialog.setValue(round(((i) / num_steps) * 100))
        # Convert DataFrame to MDF
        signals = []
        attribute_name_list = []
        time = ui.aligned_csv_time
        filtered_time = time[(time >= start_time) & (time <= end_time)]
        filtered_time = filtered_time - filtered_time[0]
        if hasattr(ui, 'csv_batt'):
            attribute_name_list.append('csv_batt')
        if hasattr(ui, 'csv_ipu'):
            attribute_name_list.append('csv_ipu')
        if hasattr(ui, 'csv_ips'):
            attribute_name_list.append('csv_ips')
        if hasattr(ui, 'csv_dcdc_high'):
            attribute_name_list.append('csv_dcdc_high')
        if hasattr(ui, 'csv_dcdc_low'):
            attribute_name_list.append('csv_dcdc_low')
        if hasattr(ui, 'csv_ecp'):
            attribute_name_list.append('csv_ecp')
        if hasattr(ui, 'csv_ptc'):
            attribute_name_list.append('csv_ptc')

        for attribute_name in attribute_name_list:
            signal_itera = generate_signals_for_csv(
                ui, attribute_name, time, filtered_time, start_time, end_time)
            signals.extend(signal_itera)

        mdf = MDF()
        mdf.append(signals)

        # Save to MDF file
        segment_file_name = f"csv_Cycle_{i+1}.mf4"
        segment_file_path = os.path.join(
            directory, segment_file_name)
        mdf.save(segment_file_path, overwrite=True)
    progressDialog.setValue(100)
    show_must_read_info("已将功率分析仪数据生成分循环的MDF数据!")


# 分循环保存转鼓数据
def loop_for_save_dyno(ui, progressDialog, directory):

    num_steps = len(ui.cycle_start_time_dyno)
    for i, (start_time, end_time) in enumerate(zip(ui.cycle_start_time_dyno, ui.cycle_end_time_dyno)):
        progressDialog.setValue(round(((i) / num_steps) * 100))
        # Convert DataFrame to MDF
        signals = []
        for index, string in enumerate(ui.dyno_data.columns):
            # Check if the substring is in the current string
            if 'TimeDiff' in string:
                # Return the index if the substring is found
                time = ui.dyno_data[ui.dyno_data.columns[index]].to_numpy().astype(
                    float)
                break
        filtered_time = time[(time >= start_time) & (time <= end_time)]
        filtered_time = filtered_time - filtered_time[0]
        # Skip the first column assumed to be time
        for column in ui.dyno_data.columns[1:]:
            values = ui.dyno_data[column].values
            filtered_value = values[(
                time >= start_time) & (time <= end_time)]
            try:
                signal = Signal(
                    samples=np.array(filtered_value, dtype=np.float64),
                    timestamps=np.array(
                        filtered_time, dtype=np.float64),
                    name=column
                )
                signals.append(signal)
            except ValueError:
                pass

        mdf = MDF()
        mdf.append(signals)

        # Save to MDF file
        segment_file_name = f"dyno_Cycle_{i+1}.mf4"
        segment_file_path = os.path.join(
            directory, segment_file_name)
        mdf.save(segment_file_path, overwrite=True)
    progressDialog.setValue(100)
    show_must_read_info("已将转鼓数据生成分循环的MDF数据!")


# 根据信号配置文件过滤信号并另存LOG数据
def save_mdf_as_filterd_signals(ui, file_path, mainWindow):
    mdf = ui.mdf
    unique_filtered_signals = []
    for signal_name in ui.filtered_signals:
        if signal_name in mdf.channels_db:
            groups_indices = mdf.channels_db[signal_name]
            # Select a random group and index if there are duplicates
            if len(groups_indices) > 1:
                group_index, channel_index = groups_indices[0]
                unique_filtered_signals.append(
                    [None, group_index, channel_index])
            else:
                group_index, channel_index = groups_indices[0]
                unique_filtered_signals.append(
                    [None, group_index, channel_index])

    filtered_mdf = mdf.filter(unique_filtered_signals)
    filtered_mdf.save(file_path, overwrite=True)
    show_must_read_info("已将MDF文件按照配置信号列表另存至目标位置!")
    mainWindow.setEnabled(True)
