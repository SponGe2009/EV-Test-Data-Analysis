import openpyxl
from openpyxl.utils import get_column_letter
from excel_info_serialization import serialize_template_details

# Function to extract details from the Excel file


def extract_excel_details(file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=False)
    details = {}  # Dictionary to store all details

    # Iterate through each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        sheet_details = {
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'cells': {},
            'row_heights': {},
            'column_widths': {},
            'merged_cells': [str(merge) for merge in sheet.merged_cells],
            'sheet_properties': {
                'tab_color': sheet.sheet_properties.tabColor
            }
        }

        # Get details for each cell
        for row in sheet.iter_rows():
            for cell in row:
                cell_details = {
                    'value': cell.value,
                    'formula': cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
                    'fill_color': {
                        "fill_type": cell.fill.fill_type,
                        "start_color": cell.fill.start_color.rgb,
                        "end_color": cell.fill.end_color.rgb,
                    },
                    'font_size': cell.font.sz,
                    'font_name': cell.font.name,
                    'border': {
                        'left': cell.border.left.style,
                        'right': cell.border.right.style,
                        'top': cell.border.top.style,
                        'bottom': cell.border.bottom.style
                    },
                    'alignment': {
                        'horizontal': cell.alignment.horizontal,
                        'vertical': cell.alignment.vertical,
                        'wrap_text': cell.alignment.wrap_text
                    }
                }
                # Assigning details to the cell's address (e.g., A1, B2)
                sheet_details['cells'][cell.coordinate] = cell_details

        # Get row heights and column widths
        for row in range(1, sheet.max_row + 1):
            sheet_details['row_heights'][row] = sheet.row_dimensions[row].height
        for col in range(1, sheet.max_column + 1):
            sheet_details['column_widths'][get_column_letter(
                col)] = sheet.column_dimensions[get_column_letter(col)].width

        # Store sheet details in the main details dictionary
        details[sheet_name] = sheet_details

    return details


if __name__ == "__main__":
    details = extract_excel_details('template_LT.xlsx')
    serialize_template_details(details, 'template_details_low_temp.json')
